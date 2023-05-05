import math
import sys
import os
import cv2
import openpyxl

from PyQt5.Qt import QUrl
from PyQt5.QtCore import QThread, Qt, QDateTime, QTimer, QRect, pyqtSignal, QPoint
from PyQt5.QtGui import QPixmap, QPainter, QPen, QImage, QColor
from PyQt5.QtWidgets import *
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer, QAbstractVideoSurface, QVideoFrame, QAbstractVideoBuffer, \
    QVideoSurfaceFormat
from PyQt5 import QtCore

from show import Ui_Form
from login import Ui_Form_login
from registered import Ui_Form_reg
from image2 import Ui_Form_image
from progressbar import Ui_Form_probar
from nec_class import myVideoSurface, MyLabel

import csv

global num_marks


# 界面一：登录界面
class video_login(QMainWindow, Ui_Form_login):
    def __init__(self, parent=None):
        super(video_login, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('篮球战术标注软件v1.0')

        self.timer = QTimer()
        self.timer.start(1000)  # 间隔1s
        # 连接信号与槽
        self.button_login.clicked.connect(self.login_test)
        self.regis.clicked.connect(self.show_2)
        self.timer.timeout.connect(self.showtime)

    # 快捷键
    def keyPressEvent(self, e):
        """
        函数说明：
        键盘为对象
        本质上该函数还是槽函数，通过键盘的输入当作信号。
        当摁下键盘就有信号输入，调用槽函数！
        """
        # 登录回车快捷键
        if e.key() == Qt.Key_Return:
            self.login_test()

    # 登录测试槽函数
    def login_test(self):
        self.user_name = self.username_line.text()
        self.password = self.password_line.text()
        if self.user_name == 'admin' and self.password == '123456':
            self.showMsg_success()
            self.show_3()
        else:
            self.book = openpyxl.load_workbook(r'password/reg_info.xlsx')
            sheet = self.book.worksheets[0]
            row = sheet.max_row
            # 从第二行开始读取
            k = 2
            while k <= row:
                templist = []
                for i in range(2):
                    cell = sheet.cell(k, i + 1)
                    templist.append(cell.value)
                if self.user_name == str(templist[0]) and self.password == str(templist[1]):
                    self.showMsg_success()
                    self.show_3()
                    break
                else:
                    k += 1
            else:
                self.showMsg_defeat()

    # 更换界面2更换槽函数
    def show_2(self):
        self.show_reg = video_reg()
        self.show_reg.show()

    # 更换界面3更换槽函数
    def show_3(self):
        # 创建界面3对象
        self.show_player = video_player()
        # 界面3 展示
        self.show_player.show()
        # 界面1 关闭
        self.close()

    # 自定义消息对话框——登陆成功 槽函数
    def showMsg_success(self):
        self.msg_success = QMessageBox()
        self.msg_success.setIcon(QMessageBox.Information)
        self.msg_success.setText("登陆成功！请等候。。。")
        self.msg_success.setWindowTitle("登陆成功！")
        self.msg_success.setStandardButtons(QMessageBox.Ok)
        self.msg_success.button(QMessageBox.Ok).animateClick(3 * 1000)  # 3秒自动关闭
        self.msg_success.exec_()
        # QMessageBox.about(self, "登陆成功", "登陆成功！")

    # 警告对话框————登陆失败 槽函数
    def showMsg_defeat(self):
        QMessageBox.warning(self, "登陆失败", "您的用户名或密码错误", QMessageBox.Retry)

    # 实时显示时间
    def showtime(self):
        time = QDateTime.currentDateTime()
        timeDisplay = time.toString("yyyy-MM-dd hh:mm:ss dddd")
        self.time.setText(timeDisplay)

    # 显示背景图片
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(self.rect(), QPixmap(r'picture/back_login.jpg'))


# 界面二：注册界面
class video_reg(QWidget, Ui_Form_reg):
    def __init__(self, parent=None):
        super(video_reg, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('注册')

        # 连接信号与槽
        self.ok.clicked.connect(self.regok)

    # 验证注册信息
    def regok(self):
        username = self.line1.text()
        pass1 = self.line2.text()
        pass2 = self.line3.text()
        if pass1 == pass2:
            templist = [username, pass1]
            self.book = openpyxl.load_workbook(r'password/reg_info.xlsx')
            sheet = self.book.worksheets[0]
            # 存在第一行覆盖bug(待解决）
            row = sheet.max_row + 1
            for i in range(len(templist)):
                sheet.cell(row, i + 1, templist[i])
            self.book.save(r'password\reg_info.xlsx')
            QMessageBox.about(self, "成功", "注册成功，信息已保存到password文件夹下！")
            self.book.close()  # 2022.11.20修改，之前是self.close()
        else:
            QMessageBox.warning(self, "失败", "两次密码不一致，请重试", QMessageBox.Retry)


# 界面三：视频截图界面
class video_player(QWidget, Ui_Form):
    def __init__(self, parent=None):
        # 装载设计好的ui界面
        super(video_player, self).__init__(parent)
        self.setupUi(self)

        # 设置窗口标题
        self.setWindowTitle("篮球战术标注软件")

        # 初始化视频播放器
        self.player = QMediaPlayer()
        self.player.setVideoOutput(self.video_play)
        # 变量初始化————>用于判断
        self.temp_frame_tuple = None
        self.appointframe = None
        self.img = myVideoSurface()
        self.cnt = 0
        self.cnt_probar = 1
        self.switch0 = 1
        self.probar_flag = 1


        self.posdir = ' '
        self.probar = video_probar()
        # 连接信号与槽
        self.button_play.clicked.connect(self.Initialize_display)
        self.video_control.clicked.connect(self.videocontrol)
        self.player.durationChanged.connect(self.getDuration)
        self.player.positionChanged.connect(self.get_sld_position)
        self.Slider.sliderMoved.connect(self.updateposition)
        self.ratechange.currentIndexChanged.connect(self.video_ratechange)
        self.fullscreen.clicked.connect(self.Fullscreen)
        self.markstart.clicked.connect(self.screenshots)  # 进行截图（未修改名）
        self.markend.clicked.connect(self.imageprocess)  # 进入图像处理（未修改名）
        self.jump.clicked.connect(self.jumpappiontframe)
        self.img.frame_available.connect(self.auto_detect)
        self.img.capture_available.connect(self.screenshots)
        self.choose.clicked.connect(self.outdir)
    # 键盘快捷键功能
    def keyPressEvent(self, e):
        """
        函数说明：
        键盘为对象
        本质上该函数还是槽函数，通过键盘的输入当作信号。
        当摁下键盘就有信号输入，调用槽函数！
        """
        # 全屏 F11
        if e.key() == Qt.Key_F11:
            self.Fullscreen()
        # 空格暂停
        elif e.key() == Qt.Key_Return:
            self.videocontrol()
    # 视频初始化操作
    def Initialize_display(self):
        # 为self.player设置播放源地址——————————在这里修改要打开的视频文件地址
        self.get_playerpath = QFileDialog.getOpenFileName(self, "选取视频文件", r"D:\DS\basketballvideo")[0]
        # 展示标签FPS,总帧数函数初始化
        def Initialize_labelshow():
            temp = cv2.VideoCapture(self.get_playerpath)
            total_frame = temp.get(cv2.CAP_PROP_FRAME_COUNT)
            fps = temp.get(cv2.CAP_PROP_FPS)

            self.fps_label.setText('FPS:{}'.format(fps))
            self.totalframe_label.setText('总帧数:{}'.format(total_frame))
            return total_frame

        self.total_frame = Initialize_labelshow()

        # 1装载播放地址并播放(解决了再次点击选择视频按钮时会出现卡顿退出的BUG）
        # 2标注EXCEL文件初始化
        if self.get_playerpath:
            # 显示正在标注视频的标题
            self.filename = os.path.basename(self.get_playerpath).rstrip('.mp4')
            self.video_title.setText('您当前正在标注视频:{}'.format(self.filename))
            # 装载播放地址并播放
            self.player.setMedia(QMediaContent(QUrl.fromLocalFile(self.get_playerpath)))
            self.player.play()

    # 视频开始或暂停
    def videocontrol(self):
        if self.player.state() == 1:
            self.player.pause()
        elif self.player.state() == 2:
            self.player.play()

    # 获取视频总时长
    def getDuration(self, d):
        self.Slider.setRange(0, d)
        self.Slider.setEnabled(True)
        self.total_timelabel(d)

    # 实现视频播放时滑动条随之滑动
    def get_sld_position(self, t):
        self.Slider.setValue(t)
        self.current_timelabel(t)

        # 如果视频播放完毕！选择重新播放或者结束
        if self.Slider.sliderPosition() == self.Slider.maximum():
            self.Slider.setValue(0)
            self.player.setPosition(0)
            rec_code = QMessageBox.question(self, '视频播放完毕', '重新标注此视频？')
            if rec_code == 65536:
                self.player.stop()
            else:
                self.player.play()

    # 实现滑动条滑动时，视频位置随之改变
    def updateposition(self, v):
        self.player.setPosition(v)

    # 显示视频播放进程文本————>修改为显示实时帧数
    def current_timelabel(self, ms):
        # self.m,self.s = divmod(ms / 1000 , 60)
        # self.current_time_label.setText('{}:{}'.format(int(self.m),int(self.s)))
        curframe = int(ms / 1000 * 25)
        self.current_time_label.setText('{}'.format(curframe))

    # 显示视频播放总长文本————>修改为显示帧数文本
    def total_timelabel(self, ms):
        # self.m, self.s = divmod(ms / 1000, 60)
        # self.total_time_label.setText('/ {}:{}'.format(int(self.m), int(self.s)))
        # 修改如下
        temp = cv2.VideoCapture(self.get_playerpath)
        total_frame = int(temp.get(cv2.CAP_PROP_FRAME_COUNT))
        self.total_time_label.setText('/ {}.'.format(total_frame))

    # 改变视频播放速率
    def video_ratechange(self, index):
        current_rate = float(self.ratechange.currentText()[1:])
        self.player.setPlaybackRate(current_rate)

    # 全屏与退出全屏
    def Fullscreen(self):
        if self.isFullScreen() == False:
            self.showFullScreen()
        else:
            self.showNormal()

    # 截图
    def screenshots(self):
        # 将要标注图片的数目保存到XLSX文件下（switch每次截图只开启一次，减少时间复杂度)
        if self.switch0 == 1:
            self.num_marks = int(self.markline.text())
            wb = openpyxl.load_workbook(r'progressbar_TempData/num_marks.xlsx')
            sh = wb["Sheet1"]  # 根据表单名称，选择sheet（表单）
            sh.cell(1, 1, self.num_marks)  # 写入数据
            wb.save(r'progressbar_TempData/num_marks.xlsx')
            wb.close()
            self.switch0 = 0  # 关闭开关
        if self.probar_flag == 1:  # 进度条初始化
            self.probar.show()
            self.probar.progressBar.setValue(0)  # 设置进度条的最小值
            self.probar.progressBar.setMaximum(self.num_marks)  # 设置进度条的最大值
            self.probar_flag = 0
        self.temp_start_frame = int(self.player.position() / 1000 * 25)
        if self.player.state() == 1 or self.player.state() == 2:
            # 先让视频暂停
            self.player.pause()
            # 设置播放窗口为myVideoSurface()获取帧数据
            self.player.setVideoOutput(self.img)  # 并不是立即执行
            # print(1)
        else:
            QMessageBox.warning(self, "标注失败", "未选择视频，请选择视频！", QMessageBox.Retry)

    # 对转换而来一帧视频流的图像进行处理，在这里进行保存图片
    def auto_detect(self, frame: QImage):
        # 首先设置在videowiget实现视频播放，使不会黑屏
        # self.player.setVideoOutput(self.video_play)
        # self.player.pause()
        if frame:
            # 保存图片demo
            # frame.save(r'DS_image\image\{}.jpg'.format(
            #     '{}'.format(self.filename.split('_')[1]) + '_' + str(self.temp_start_frame)))
            frame.save(r'{}\{}.jpg'.format(
                self.posdir,'{}'.format(self.filename.split('_')[1]) + '_' + str(self.temp_start_frame)))
            self.probar.progressBar.setValue(self.cnt_probar)
            self.cnt_probar += 1
            print('Done!:', self.temp_start_frame)
            # QMessageBox.about(self, "成功", "{}截图成功！".format(self.temp_start_frame))
            self.player.play()
            self.cnt += 1
            if self.cnt == self.num_marks:
                self.cnt = 0
                self.cnt_probar = 1
                self.probar_flag = 1
                self.switch0 = 1
                self.probar.close()
                self.player.setVideoOutput(self.video_play)
                self.player.pause()
                QMessageBox.about(self, "成功", "截图完毕！")
                # 此处可添加让视频播放 self.player.play() 看用户需求
        else:
            return

    # 进入图像处理界面
    def imageprocess(self):
        # if self.player.state() == 1 or self.player.state() == 2:
        # 先让视频暂停
        self.player.pause()
        # 弹出图像处理界面
        self.video_image = video_image()
        self.video_image.show()

    # 跳转到指定帧数
    def jumpappiontframe(self):
        temp = cv2.VideoCapture(self.get_playerpath)
        total_frame = temp.get(cv2.CAP_PROP_FRAME_COUNT)
        self.appointframe = self.assginframe.text()
        if self.appointframe.isdigit() and int(self.appointframe) <= total_frame:
            self.appointframe = int(self.assginframe.text())
            appointtime = int(self.appointframe / 25 * 1000)
            self.player.setPosition(appointtime)
        elif int(self.appointframe) > total_frame:
            QMessageBox.warning(self, "失败", "请输入正确帧数！", QMessageBox.Retry)
    # 输出图像保存目标地址
    def outdir(self):
        self.posdir = QFileDialog.getExistingDirectory \
            (self, "选取文件", r"C:\Users\20561\PycharmProjects\pythonProject\Basketballlabelingsoftware\DS_image")
        self.label.setText('输出的图像位置:{}'.format(self.posdir))
    # 更换背景
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(self.rect(), QPixmap(r'picture/back_show.jpg'))


# 界面四：图像处理界面
class video_image(QWidget, Ui_Form_image):
    # 自动切换标志
    flag2 = True

    def __init__(self, parent=None):
        super(video_image, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('图像处理')
        self.get_playerpath = ''
        self.get_path = QFileDialog.getExistingDirectory\
            (self, "选取文件", r"C:\Users\20561\PycharmProjects\pythonProject\Basketballlabelingsoftware\DS_image")
        self.image_widget = ImageWidget(self, dir=self.get_path, col=2, w=1800, suit=0)
        self.image_widget.move(5, 60)
        # 连接信号与槽
        self.pb_pre.clicked.connect(lambda: self.image_widget.turn_page(-1))
        self.pb_next.clicked.connect(lambda: self.image_widget.turn_page(1))  # 图像列表翻页
        self.skip.clicked.connect(lambda: self.image_widget.skip_page(int(self.skipline.text())))  # 跳转
        self.image_widget.signal_order.connect(self.change_path)
        self.image_widget.signal_page.connect(self.change_page)

        self.undo1.clicked.connect((lambda: self.undo(self.image_widget.group_num * 2 - 2)))
        self.undo2.clicked.connect((lambda: self.undo(self.image_widget.group_num * 2 - 1)))

        self.out1.clicked.connect((lambda: self.saveimg(self.image_widget.group_num * 2 - 2)))
        self.out2.clicked.connect((lambda: self.saveimg(self.image_widget.group_num * 2 - 1)))
        # 自动保存槽函数
        self.image_widget.signal1.connect(self.saveimg1)
        self.image_widget.signal2.connect(self.saveimg1)

        self.flashbutton.clicked.connect(self.flash)
        self.choose.clicked.connect(self.chooseimg)
        self.comb = [self.comb1, self.comb2, self.comb3, self.comb4, self.comb5, self.comb6, self.comb7, self.comb8,
                     self.comb9, self.comb10]

        self.skip_s = [0 for i in range(10)]
        self.flag1 = False
        self.id = (
            ("1",), ("2",), ("3",), ("4",), ("5",), ("a",), ("b",), ("c",), ("d",),
            ("e",))  # 定义边缘框id
        self.action_id = {'盖帽': 1, '罚篮': 2, '传球': 3, '防守': 4, '进攻': 5, '投篮': 6,
                          '站位': 7, '挡拆': 8, '跑位': 9}
        try:
            self.color_list = list(self.image_widget.Label[0].color_dict.values()) #没有图片这段语句会报错
        except Exception as err:
            QMessageBox.about(self, "异常", "无图片，请导入图片后重试")
    def chooseimg(self):
        self.last_files = self.image_widget.list_files
        self.get_playerpath = QFileDialog.getExistingDirectory(self, "选取文件", r"DS_image")
        self.image_widget.list_files = []
        # 选择退出处理
        if self.get_playerpath != '':
            self.image_widget.get_files(self.get_playerpath)
        else:
            return
        # 文件夹为空处理
        if self.image_widget.list_files == []:
            self.image_widget.list_files = self.last_files
            self.image_widget.show_images_list()
            QMessageBox.about(self, "异常", "该文件夹无图片，请导入图片后重试")
        else:
            self.image_widget.show_images_list()
    def keyPressEvent(self, e):
        """
        函数说明：
        键盘为对象
        本质上该函数还是槽函数，通过键盘的输入当作信号。
        当摁下键盘就有信号输入，调用槽函数！
        """
        # 切换不保存图片模式，只单纯的重复锚框用来获取数据
        if e.key() == Qt.Key_F1:
            self.change_mode1()
        if e.key() == Qt.Key_F2:
            self.change_mode2()

    def change_mode1(self):
        if self.flag1:
            self.flag1 = False
            self.modeLabel.setText("不截图模式")
        else:
            self.flag1 = True
            self.modeLabel.setText("截图模式")

    def change_mode2(self):
        if self.flag2:
            self.flag2 = False
            self.image_widget.flag = False
            self.label.setText("手动导出模式")
        else:
            self.flag2 = True
            self.image_widget.flag = True
            self.label.setText("自动导出模式")

    def change_path(self, path):
        self.dir.setText(path)
    # 跳转页
    def change_page(self, index):
        self.page.setText(f"第{index}页")

    # 保存标注结果（手动）
    def saveimg(self, index):  # 按Labelimg实现每一页保存一次！摒弃多页保存
        if self.flag2 == False:
            # grabWindow 继承 Qscreen: 创建并返回一个像素映射
            # 窗口系统标识符()可以使用QWidget.winId()函数检索
            if self.image_widget.Label[index].rect_position:
                filename = os.path.basename(self.image_widget.list_files[index]).strip('.jpg')
                si = filename.index('_')
                video_name = filename[0:si]
                frame_id = filename[si + 1:]
                with open(r"DS_csv\tempdata.csv", mode="a", encoding="utf-8-sig", newline="") as f:
                    # 基于打开的文件，创建 csv.writer 实例
                    writer = csv.writer(f)
                    if self.flag1:
                        QApplication.primaryScreen().grabWindow(self.image_widget.Label[index].winId()).save(
                            self.image_widget.list_files[index])  # 保存当前的label界面
                    for ind in range(10):
                        # ——————————————————————————————————————————————该判断语句实现两个功能—————————————————————————————————————————
                        if self.image_widget.Label[index].color_dict[
                            self.image_widget.Label[index].color_dict_keys[ind]] != None \
                                and self.image_widget.Label[index].rect_position[ind] != 0:
                            actionid = self.action_id[self.comb[ind].currentText()]
                            self.temp_attribute = (video_name, frame_id) + \
                                                  self.image_widget.Label[
                                                      index].rect_attribute[
                                                      ind] + \
                                                  (actionid,) + self.id[ind]
                            writer.writerow(self.temp_attribute)
                    QMessageBox.about(self, "成功", "导出成功，请查看csv文件")
                    print('{}手动导出成功'.format(frame_id))
        else:
            QMessageBox.about(self, "失败", "手动导出模式禁用！")
    # 保存标注结果（自动）
    def saveimg1(self, index):  # 按Labelimg实现每一页保存一次！摒弃多页保存
        # grabWindow 继承 Qscreen: 创建并返回一个像素映射
        # 窗口系统标识符()可以使用QWidget.winId()函数检索
        if self.image_widget.Label[index].rect_position:
            filename = os.path.basename(self.image_widget.list_files[index]).strip('.jpg')
            si = filename.index('_')
            video_name = filename[0:si]
            frame_id = filename[si + 1:]
            with open(r"DS_csv\tempdata.csv", mode="a", encoding="utf-8-sig", newline="") as f:
                # 基于打开的文件，创建 csv.writer 实例
                writer = csv.writer(f)
                if self.flag1:
                    QApplication.primaryScreen().grabWindow(self.image_widget.Label[index].winId()).save(
                        self.image_widget.list_files[index])  # 保存当前的label界面
                for ind in range(10):
                    # ——————————————————————————————————————————————该判断语句实现两个功能—————————————————————————————————————————
                    if self.image_widget.Label[index].color_dict[
                        self.image_widget.Label[index].color_dict_keys[ind]] != None \
                            and self.image_widget.Label[index].rect_position[ind] != 0:
                        actionid = self.action_id[self.comb[ind].currentText()]
                        self.temp_attribute = (video_name, frame_id) + \
                                              self.image_widget.Label[
                                                  index].rect_attribute[
                                                  ind] + \
                                              (actionid,) + self.id[ind]
                        writer.writerow(self.temp_attribute)
                print('{}自动导出成功'.format(frame_id))

    # 撤销
    def undo(self, index):
        # 撤销矩形框
        self.image_widget.Label[index].rectangles.pop(-1)
        # 属性值归0
        self.image_widget.Label[index].rect_position[self.image_widget.Label[index].attrcnt - 1] = 0
        self.image_widget.Label[index].rect_attribute[self.image_widget.Label[index].attrcnt - 1] = 0
        # 计数减一
        self.image_widget.Label[index].attrcnt -= 1

        QMessageBox.about(self, "成功", "撤销成功!")

    # 刷新矩形框
    def flash(self):
        cnt = 0

        color_dict = {"red": Qt.red, "black": Qt.black, "yellow": Qt.yellow, "blue": Qt.blue, "green": Qt.green,
                      "gray": Qt.gray,
                      "white": Qt.white, "cyan": Qt.cyan, "magenta": Qt.magenta,
                      "orange": QColor(255, 140, 0)}  # 防止溢出冲突
        self.flash_list = [self.checkBox1, self.checkBox2, self.checkBox3, self.checkBox4, self.checkBox5,
                           self.checkBox6, self.checkBox7, self.checkBox8, self.checkBox9, self.checkBox10]

        for keys, values in color_dict.items():
            if not self.flash_list[cnt].isChecked():
                self.image_widget.Label[0].color_dict[keys] = self.color_list[cnt]
            else:
                self.image_widget.Label[0].color_dict[keys] = None
                self.image_widget.Label[0].color_flag[cnt] = 0
            cnt += 1
        QMessageBox.about(self, "成功", "刷新成功!")
    # 显示背景图片
    # def paintEvent(self, event):
    #     painter = QPainter(self)
    #     painter.drawPixmap(self.rect(), QPixmap(r'picture/back.png'))


# 自定义图片控件 界面四的内容
class ImageWidget(QWidget):
    group_num = 1  # 图像列表当前组数（页数）
    #list_files = []  # 图像文件路径集
    signal_order = pyqtSignal(str)  # 图像项目信号
    signal_page = pyqtSignal(int)  # 页数信号
    Label = [0 for i in range(10000)]  # 存放容量为10000个的Label

    signal1 = pyqtSignal(int)
    signal2 = pyqtSignal(int)

    def __init__(self, parent=None, dir='./', col=1, w=10, h=500, suit=0):
        super(ImageWidget, self).__init__(parent)
        self.list_files = []  # 图像文件路径集
        self.get_files(dir)
        self.skipsignal = False  # 跳转标志
        self.col = col  # 一页显示几张照片
        self.w = w
        self.suit = suit
        if h == None:
            self.h = int(self.w / self.col)
        else:
            self.h = h
        self.setFixedSize(self.w, self.h)
        self.hbox = QHBoxLayout(self)
        self.hbox.setContentsMargins(0, 0, 0, 0)  # 设置框边距
        self.show_images_list()  # 初次加载图像列表
        self.flag = True

    def get_files(self, dir):  # 储存需加载的所有图像路径
        for file in os.listdir(path=dir):
            if file.endswith('jpg') or file.endswith('png'):
                self.list_files.append(dir + "/" + file)
        self.list_files.sort(key = lambda x:int(x.split('_')[2].split('.')[0]))# 字符串按位比较ascii码，要转int型

    def turn_page(self, num):  # 图像列表翻页
        ##########翻页同时保存前两张图片数据###############
        if self.flag == True:
            self.signal1.emit(self.group_num * 2 - 2)
            self.signal2.emit(self.group_num * 2 - 1)
            # QMessageBox.about(self, "成功", "ok!")
        flag = len(self.list_files)
        if self.group_num == 1 and num == -1:  # 到首页时停止上翻
            QMessageBox.about(self, "Remind", "这是第一张图像!")
        elif (self.group_num == math.ceil(flag / self.col) and num == 1) or flag == 0:  # 到末页页时停止下翻
            QMessageBox.about(self, "Remind", "没有更多图像啦")
        else:
            self.group_num += num  # 翻页
        self.signal_page.emit(self.group_num)
        self.show_images_list()  # 重新加载图像列表

    def skip_page(self, num):
        flag = len(self.list_files)
        if self.group_num == 1 and num == -1:  # 到首页时停止上翻
            QMessageBox.about(self, "Remind", "这是第一张图像!")
        elif (self.group_num == math.ceil(flag / self.col) and num == 1) or flag == 0:  # 到末页页时停止下翻
            QMessageBox.about(self, "Remind", "没有更多图像啦")
        else:
            self.group_num = num  # 翻页
            self.skipsignal = True
        self.signal_page.emit(self.group_num)
        self.show_images_list()  # 重新加载图像列表

    def show_images_list(self):  # 加载图像列表
        for i in range(self.hbox.count() - 1):  # 每次加载先清空内容，避免layout里堆积label
            self.hbox.itemAt(i).widget().deleteLater()
        # 设置分段显示图像，每col个一段
        group_num = self.group_num
        start = 0
        end = self.col
        if group_num > 1:
            start = self.col * (group_num - 1)
            end = self.col * group_num
        count = 0  # 记录当前页载入的label数
        width = int(self.w / (self.col + 1))  # 自定义label宽度
        height = self.h  # 自定义label高度
        for index, path in enumerate(self.list_files):  # group_num = 1 则加载前col个，以此类推
            # 通过if elif来控制col
            if index < start:
                continue
            elif index == end:
                break
            # 按路径读取成QPixmap格式的图像，根据适应方式调整尺寸
            if self.suit == 0:
                if index == 0:
                    pix = QPixmap(path).scaled(600 - 2 * self.col, height - 4)
                elif index == 1:
                    pix = QPixmap(path).scaled(600 - 2 * self.col, height - 4)
                else:
                    pix = QPixmap(path).scaled(width - self.col, height - 4)
            elif self.suit == 1:
                pix = QPixmap(path)
                pix = QPixmap(path).scaled(int(pix.width() * height / pix.height()) - 2 * self.col, height - 4)
            elif self.suit == 2:
                pix = QPixmap(path)
                pix = QPixmap(path).scaled(width - 2 * self.col, int(pix.height() * width / pix.width()) - 4)
            label = MyLabel(index)
            self.Label[index] = label  # 把Label保存在列表中
            label.setPixmap(pix)  # 加载图片
            self.hbox.addWidget(label)  # 在水平布局中添加自定义label
            label.signal_order.connect(self.choose_image)  # 绑定自定义label点击信号 2
            count += 1
            if not count == self.col:
                for i in range(self.col - count):
                    label = QLabel()
                    self.hbox.addWidget(label)  # 在水平布局中添加空label补位

    def choose_image(self, index):  # 选择图像
        self.signal_order.emit(self.list_files[index])  # 3 绑定自己类的signal_order


# 进度条
class video_probar(QWidget, Ui_Form_probar):
    def __init__(self, parent=None):
        super(video_probar, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle('进度')
